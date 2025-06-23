#!/usr/bin/env python3
"""
GitHub Commit Calculator
Calculate unique user commits for multiple repositories and time ranges.
"""

import os
import sys
import json
from collections import defaultdict
from datetime import datetime, timezone
from typing import Dict, List, Optional, Any
import click
from github import Github, GithubException
from github.Repository import Repository
from rich.console import Console
from rich.table import Table
from rich.progress import Progress, SpinnerColumn, TextColumn
from dotenv import load_dotenv
import yaml
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Load environment variables
load_dotenv()

console = Console()


class GitHubCommitCalculator:
    """Calculate unique user commits across all branches in a GitHub repository."""

    def __init__(self, token: str, org_name: str, repo_name: str, exclude_merge_commits: bool = True):
        self.github = Github(token)
        self.org_name = org_name
        self.repo_name = repo_name
        self.exclude_merge_commits = exclude_merge_commits
        self.repo: Optional[Repository] = None
        self._authenticate()

    def _authenticate(self):
        try:
            org = self.github.get_organization(self.org_name)
            self.repo = org.get_repo(self.repo_name)
        except GithubException as e:
            if e.status == 404:
                console.print(f"❌ Repository {self.org_name}/{self.repo_name} not found", style="red")
            elif e.status == 401:
                console.print("❌ Authentication failed. Please check your GitHub token.", style="red")
            else:
                console.print(f"❌ Error authenticating with {self.org_name}/{self.repo_name}: {e}", style="red")
            raise e

    def get_all_branches(self) -> List[str]:
        if not self.repo:
            return []
        branches = []
        try:
            for branch in self.repo.get_branches():
                branches.append(branch.name)
            console.print(f"📋 Found {len(branches)} branches for {self.repo_name}", style="blue")
            return branches
        except GithubException as e:
            console.print(f"❌ Error fetching branches for {self.repo_name}: {e}", style="red")
            return []

    def is_merge_commit(self, commit) -> bool:
        if len(commit.parents) > 1:
            return True
        message = commit.commit.message.lower()
        merge_indicators = ['merge pull request', 'merge branch', 'merge remote', 'merge from', 'merge into', 'merge:', 'merged', 'merging']
        return any(indicator in message for indicator in merge_indicators)

    def get_commits_for_branch(self, branch_name: str, start_date: Optional[datetime] = None, end_date: Optional[datetime] = None) -> List[Dict]:
        if not self.repo:
            return []
        commits = []
        merge_commits_excluded = 0
        params: Dict[str, Any] = {'sha': branch_name}
        if start_date:
            params['since'] = start_date
        if end_date:
            params['until'] = end_date
        
        try:
            for commit in self.repo.get_commits(**params):
                if self.exclude_merge_commits and self.is_merge_commit(commit):
                    merge_commits_excluded += 1
                    continue
                
                commits.append({
                    'author': commit.author.login if commit.author else 'Unknown',
                    'date': commit.commit.author.date,
                    'message': commit.commit.message,
                    'tree_sha': commit.commit.tree.sha
                })
            
            if self.exclude_merge_commits and merge_commits_excluded > 0:
                console.print(f"📝 Branch {branch_name}: Excluded {merge_commits_excluded} merge commits", style="dim")
            
            return commits
        except GithubException as e:
            console.print(f"⚠️  Error fetching commits for branch {branch_name}: {e}", style="yellow")
            return []

    def calculate_commits(self, branches: Optional[List[str]] = None, start_date: Optional[datetime] = None, end_date: Optional[datetime] = None) -> Dict:
        if branches is None:
            branches = self.get_all_branches()
        
        user_stats: Dict[str, Dict[str, Any]] = defaultdict(lambda: {
            'unique_commits': 0, 'branches': set(), 'unique_commits_by_branch': defaultdict(int),
            'first_commit': None, 'last_commit': None
        })
        
        seen_commits = set()
        
        with Progress(SpinnerColumn(), TextColumn("[progress.description]{task.description}"), console=console) as progress:
            task = progress.add_task("Analyzing commits...", total=len(branches))
            for branch in branches:
                progress.update(task, description=f"Processing branch: {branch}")
                commits = self.get_commits_for_branch(branch, start_date, end_date)
                
                for commit in commits:
                    author = commit['author']
                    commit_key = (author, commit['message'], commit['tree_sha'])
                    
                    if commit_key not in seen_commits:
                        seen_commits.add(commit_key)
                        stats = user_stats[author]
                        stats['unique_commits'] += 1
                        stats['branches'].add(branch)
                        stats['unique_commits_by_branch'][branch] += 1
                        
                        commit_date = commit['date']
                        if stats['first_commit'] is None or commit_date < stats['first_commit']:
                            stats['first_commit'] = commit_date
                        if stats['last_commit'] is None or commit_date > stats['last_commit']:
                            stats['last_commit'] = commit_date
                progress.advance(task)

        for user in user_stats:
            user_stats[user]['branches'] = list(user_stats[user]['branches'])
            user_stats[user]['unique_commits_by_branch'] = dict(user_stats[user]['unique_commits_by_branch'])
        
        return {'unique_commits': len(seen_commits), 'total_branches': len(branches), 'user_stats': dict(user_stats)}

    def display_results(self, stats: Dict, detailed: bool, time_range_name: Optional[str] = None):
        header = time_range_name if time_range_name else "All Time"
        console.print(f"\n--- Results for {header} ---", style="bold cyan")
        
        summary_table = Table(title=f"Repository Summary ({header})")
        summary_table.add_column("Metric", style="cyan")
        summary_table.add_column("Value", style="magenta")
        summary_table.add_row("Unique Commits", str(stats['unique_commits']))
        summary_table.add_row("Total Branches Analyzed", str(stats['total_branches']))
        summary_table.add_row("Active Contributors", str(len(stats['user_stats'])))
        console.print(summary_table)

        sorted_users = sorted(stats['user_stats'].items(), key=lambda x: x[1]['unique_commits'], reverse=True)
        user_table = Table(title=f"User Ranking ({header})")
        user_table.add_column("Rank", style="cyan", justify="center")
        user_table.add_column("User", style="green")
        user_table.add_column("Unique Commits", style="magenta", justify="center")
        user_table.add_column("Branches", style="yellow", justify="center")
        user_table.add_column("First Commit", style="blue")
        user_table.add_column("Last Commit", style="blue")
        
        for rank, (user, data) in enumerate(sorted_users, 1):
            first = data['first_commit'].strftime('%Y-%m-%d') if data['first_commit'] else 'N/A'
            last = data['last_commit'].strftime('%Y-%m-%d') if data['last_commit'] else 'N/A'
            user_table.add_row(str(rank), user, str(data['unique_commits']), str(len(data['branches'])), first, last)
        console.print(user_table)
        
        if detailed:
            for user, data in sorted_users:
                if data['unique_commits'] == 0: continue
                console.print(f"\n👤 {user} ({data['unique_commits']} unique commits)", style="bold green")
                branch_table = Table(title=f"Branch Activity for {user} ({header})")
                branch_table.add_column("Branch", style="cyan")
                branch_table.add_column("Unique Commits", style="magenta", justify="center")
                branch_table.add_column("Percentage", style="yellow", justify="center")
                
                total_commits = data['unique_commits']
                for branch, commits in sorted(data['unique_commits_by_branch'].items(), key=lambda x: x[1], reverse=True):
                    percentage = (commits / total_commits) * 100 if total_commits > 0 else 0
                    branch_table.add_row(branch, str(commits), f"{percentage:.1f}%")
                console.print(branch_table)

def save_to_excel(all_results: Dict, filename: str, detailed: bool):
    """Saves the analysis results to an Excel file."""
    workbook = openpyxl.Workbook()
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')

    for repo_name, repo_data in all_results.items():
        sheet_name = repo_name[:31]
        ws = workbook.create_sheet(title=sheet_name)
        
        ws.cell(row=1, column=1, value=f"Analysis for Repository: {repo_name}").font = Font(bold=True, size=16)
        
        current_row = 3
        
        for time_range, stats in repo_data.items():
            ws.cell(row=current_row, column=1, value=f"Time Range: {time_range}").font = Font(bold=True, size=14)
            current_row += 1
            
            headers = ["Rank", "User", "Unique Commits", "Branches", "First Commit", "Last Commit"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

            sorted_users = sorted(stats['user_stats'].items(), key=lambda x: x[1]['unique_commits'], reverse=True)
            
            for rank, (user, data) in enumerate(sorted_users, 1):
                first = data['first_commit'].strftime('%Y-%m-%d') if data['first_commit'] else 'N/A'
                last = data['last_commit'].strftime('%Y-%m-%d') if data['last_commit'] else 'N/A'
                
                ws.cell(row=current_row + rank, column=1, value=rank).alignment = center_align
                ws.cell(row=current_row + rank, column=2, value=user)
                ws.cell(row=current_row + rank, column=3, value=data['unique_commits']).alignment = center_align
                ws.cell(row=current_row + rank, column=4, value=len(data['branches'])).alignment = center_align
                ws.cell(row=current_row + rank, column=5, value=first).alignment = center_align
                ws.cell(row=current_row + rank, column=6, value=last).alignment = center_align
            
            current_row += len(sorted_users) + 2
        
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    try:
        workbook.save(filename)
        console.print(f"✅ Successfully saved Excel report to [bold]{filename}[/bold]", style="green")
    except Exception as e:
        console.print(f"❌ Error saving Excel file: {e}", style="red")

def load_config(config_path: str = "config.yaml") -> dict:
    if not os.path.exists(config_path):
        console.print(f"⚠️  Config file {config_path} not found.", style="yellow")
        return {}
    with open(config_path, 'r') as f:
        return yaml.safe_load(f) or {}

@click.command()
@click.option('--config', '-c', default='config.yaml', help='Path to config YAML file.')
@click.option('--org', '-o', help='GitHub organization name.')
@click.option('--repo', '-r', 'repos', multiple=True, help='Repository name (can be used multiple times).')
@click.option('--token', '-t', envvar='GITHUB_TOKEN', help='GitHub personal access token.')
@click.option('--branches', '-b', multiple=True, help='Specific branches to analyze.')
@click.option('--detailed', '-d', is_flag=True, help='Show detailed branch statistics for each user.')
@click.option('--include-merge-commits', is_flag=True, help='Include merge commits in the analysis.')
@click.option('--output', '-O', type=click.Path(), help='Output file for results (JSON format).')
@click.option('--output-excel', type=click.Path(), help='Output file for results (Excel format).')
def main(config, org, repos, token, branches, detailed, include_merge_commits, output, output_excel):
    """Calculate unique user commits for multiple repositories and time ranges."""
    config_data = load_config(config)
    
    org = org or config_data.get('organization')
    repositories = repos if repos else config_data.get('repositories', [])
    token = token or os.getenv('GITHUB_TOKEN')
    
    if not token or not org or not repositories:
        console.print("❌ Token, organization, and at least one repository are required.", style="red")
        sys.exit(1)
    
    overall_stats = {}

    for repo_name in repositories:
        console.print(f"\n\n{'='*25}\n ANALYZING REPOSITORY: {org}/{repo_name} \n{'='*25}", style="bold yellow on_black")
        try:
            calculator = GitHubCommitCalculator(token, org, repo_name, not include_merge_commits)
            branch_list = list(branches) if branches else config_data.get('branches')
            if not branch_list:
                branch_list = None
            
            time_ranges = config_data.get('time_ranges')
            repo_stats = {}

            if not time_ranges:
                console.print("🚀 No time ranges defined. Analyzing entire history.", style="bold green")
                stats = calculator.calculate_commits(branch_list)
                calculator.display_results(stats, detailed, "All Time")
                repo_stats["All Time"] = stats
            else:
                console.print(f"🚀 Analyzing {len(time_ranges)} defined time range(s).", style="bold green")
                for tr in time_ranges:
                    name = tr.get('name', 'Unnamed Range')
                    start_date = datetime.fromisoformat(tr['start_date']).replace(tzinfo=timezone.utc) if 'start_date' in tr else None
                    end_date = datetime.fromisoformat(tr['end_date']).replace(tzinfo=timezone.utc) if 'end_date' in tr else None
                    
                    stats = calculator.calculate_commits(branch_list, start_date, end_date)
                    calculator.display_results(stats, detailed, name)
                    repo_stats[name] = stats
            
            overall_stats[repo_name] = repo_stats

        except GithubException:
             console.print(f"--- Skipping repository {repo_name} due to authentication/not found error ---", style="bold red")
             continue
        except Exception as e:
            console.print(f"❌ An unexpected error occurred while analyzing {repo_name}: {e}", style="red")
            continue

    if output:
        # Create a deep copy for JSON serialization to avoid modifying the original data
        json_output_stats = json.loads(json.dumps(overall_stats, default=str))
        with open(output, 'w') as f:
            json.dump(json_output_stats, f, indent=4)
        console.print(f"✅ Successfully saved results to [bold]{output}[/bold]", style="green")

    if output_excel:
        save_to_excel(overall_stats, output_excel, detailed)

if __name__ == '__main__':
    main()